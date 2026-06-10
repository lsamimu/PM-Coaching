"""Build the Excel deliverables from the crawled jobs.

Reads .tmp/jobs.json and writes TWO separate workbooks into output/:
  - pm_internships_<date>.xlsx  (sheet "Internships")
  - pm_fulltime_<date>.xlsx     (sheet "Full-Time")

Each row: Role | Company | Experience Level | Apply Link | Job Summary |
          Application Deadline | Date Posted | Location | Source

Run:  python execution/build_excel.py
"""
import json
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

import config

COLUMNS = [
    ("Role", 34),
    ("Company", 24),
    ("Experience Level", 20),
    ("Apply Link", 40),
    ("Job Summary", 70),
    ("Application Deadline", 20),
    ("Date Posted", 14),
    ("Location", 22),
    ("Source", 18),
]

HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
WRAP = Alignment(wrap_text=True, vertical="top")


def _write_sheet(ws, jobs):
    # Header row.
    for col_idx, (name, width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center", horizontal="center")
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 22

    # Data rows.
    for row_idx, job in enumerate(jobs, start=2):
        values = [
            job.get("role", ""),
            job.get("company", ""),
            job.get("experience_level", ""),
            job.get("link", ""),
            job.get("summary", ""),
            job.get("deadline", "Not specified"),
            job.get("posted", ""),
            job.get("location", ""),
            job.get("source", ""),
        ]
        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = WRAP
            # Make the apply link clickable.
            if col_idx == 4 and value:
                cell.hyperlink = value
                cell.font = Font(color="0563C1", underline="single")

    # Add an autofilter over the populated range.
    last_col = get_column_letter(len(COLUMNS))
    last_row = max(len(jobs) + 1, 1)
    ws.auto_filter.ref = f"A1:{last_col}{last_row}"


def _build_one(jobs, sheet_title, filename):
    """Write a single-sheet workbook for one category and return its path."""
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title
    _write_sheet(ws, jobs)
    out_path = config.OUTPUT_DIR / filename
    wb.save(out_path)
    return out_path


def build(jobs_data=None):
    """Build two separate workbooks (internships + full-time).

    Returns a list of the saved file paths.
    """
    config.ensure_dirs()
    if jobs_data is None:
        jobs_data = json.loads(config.JOBS_JSON.read_text())

    internships = jobs_data.get("internships", [])
    fulltime = jobs_data.get("fulltime", [])
    date_str = datetime.now().strftime("%Y-%m-%d")

    intern_path = _build_one(internships, "Internships", f"pm_internships_{date_str}.xlsx")
    fulltime_path = _build_one(fulltime, "Full-Time", f"pm_fulltime_{date_str}.xlsx")

    print(f"Workbook saved -> {intern_path} (Internships: {len(internships)})")
    print(f"Workbook saved -> {fulltime_path} (Full-Time: {len(fulltime)})")
    return [intern_path, fulltime_path]


if __name__ == "__main__":
    build()
